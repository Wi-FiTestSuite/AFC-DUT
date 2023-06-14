from openpyxl import load_workbook
import json
import re

vectors = {"vectors": []}

wb = load_workbook("AFC Device (DUT) Compliance Test Vectors v1.0.xlsx")

ws = wb.worksheets[4]

for row in range(7, 25):
    vec = {}
    tc_id = {}
    response = {}
    tc_id_row = row
    phase = 0

    if ws[f'A{row}'].value and not ws[f'A{row + 1}'].value:
       phase = 1
    elif not ws[f'A{row}'].value:
       phase = 2
       tc_id_row = row - 1

    # Test Case ID as index
    tc_id["unitUnderTest"] = ws[f'A{tc_id_row}'].value
    tc_id["purpose"] = ws[f'B{tc_id_row}'].value
    tc_id["testVector"] = ws[f'C{tc_id_row}'].value
    if phase:
        tc_id["phase"] = phase
        filename = f"{ws[f'A{tc_id_row}'].value}_{ws[f'B{tc_id_row}'].value}_{ws[f'C{tc_id_row}'].value}_phase{phase}.json"
    else:
        filename = f"{ws[f'A{tc_id_row}'].value}_{ws[f'B{tc_id_row}'].value}_{ws[f'C{tc_id_row}'].value}.json"
    vec["testCaseID"] = tc_id
    
    ###
    vec["description"] = ws[f'D{tc_id_row}'].value
    response["availableSpectrumInquiryResponses"] = []
    vec["responses"] = response
    InquiryResponse = {}
    response["availableSpectrumInquiryResponses"].append(InquiryResponse)
    InquiryResponse["response"] = {
                                   "responseCode": ws[f'EW{row}'].value,
                                   "shortDescription": ws[f'EX{row}'].value
                                  }

    # availableFrequencyInfo
    # r6: Column H ~ EF
    data = [ws.cell(row=row,column=i).value for i in range(8,137)]
    print(f'row - {row}')
    print(data)
    for i in range(0, len(data), 3):
        if data[i]:
            if not InquiryResponse.get("availableFrequencyInfo"):
                InquiryResponse["availableFrequencyInfo"] = []
            if data[i+2] != 'NULL':
                print(data[i+2])
                InquiryResponse["availableFrequencyInfo"].append({
                                    "frequencyRange": {
                                        "highFrequency": data[i+1],
                                        "lowFrequency": data[i]
                                    },
                                    "maxPsd": data[i+2]
                                })    


    # availableChannelInfo
    # cell EG ~ EU
    data = [ws.cell(row=row,column=i).value for i in range(137,152)]
    print(data)
    for i in range(0, len(data), 3):
        if data[i]:
            print(f'gobalOperatingClass {data[i]}')
            if not InquiryResponse.get("availableChannelInfo"):
                InquiryResponse["availableChannelInfo"] = []
            channelCfi = re.findall('([^,"\t\n ]+)', str(data[i+1]))
            print(f"channelCfi {channelCfi} len {len(channelCfi)}")
            maxEirp = re.findall('([^,"\t\n ]+)', str(data[i+2]))
            print(f"maxEirp    {maxEirp} len {len(maxEirp)}")
            if len(maxEirp) == 1 and (maxEirp[0] == 'NULL' or maxEirp[0] == 'None'):
                continue
            InquiryResponse["availableChannelInfo"].append({
                                 "channelCfi":  [int(item) for idx, item in enumerate(channelCfi) if item and item != "NULL" and len(maxEirp) > idx and maxEirp[idx] and maxEirp[idx] != 'NULL' ],
                                "globalOperatingClass": data[i],
                                "maxEirp": [float(item) for item in maxEirp if item and item != "NULL"]
                            })    
    #vec["row"] = row # debug
    #vec["tc_id_row"] = tc_id_row # debug
    with open(filename, "w") as f:
        json.dump(vec, f, indent=4)

